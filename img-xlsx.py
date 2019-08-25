from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from functools import partial
import sys
import argparse

def rgb_to_xls_hex(rgb_tuple, image_mode='RGB'):
    if image_mode == 'RGB':
        r, g, b = rgb_tuple
    elif image_mode == 'RGBA':
        # Ignore alpha channel in images that have one.
        r, g, b, _ = rgb_tuple
    return f'{r:02x}{g:02x}{b:02x}'


def handle_arguments():
    parser = argparse.ArgumentParser(description='Convert an image file to an Excel spreadsheet. I\'m sorry.')
    parser.add_argument('--size', dest='size', type=int, default=64,
                        help='The number of cells for the largest dimension of the image. '
                             'Defaults to 64. Up to 512 works well for landscape images, up to 256 '
                             'for portrait images.')
    parser.add_argument('--quantize', dest='quantize', metavar='NUM_COLORS', type=int, default=0,
                        help='Quantize the image (i.e. set an upper bound on the number of colors). '
                             'Max 255.')
    parser.add_argument('image', metavar='FILENAME', type=str, 
                        help='The image file to turn into an Excel spreadsheet. JPGs and PNGs work well.')
    parser.add_argument('xlsx', metavar='FILENAME', type=str,
                        help='The output filename. Should end in .xlsx')

    args = parser.parse_args()
    return args


def convert(args):
    im = Image.open(args.image)

    maxsize = (args.size, args.size)
    im.thumbnail(maxsize)

    if args.quantize > 0 and args.quantize < 256:
        quantized = im.quantize(colors=args.quantize)
        im = quantized

    if im.mode in ['P', 'L']:
        image = im.convert("RGB")
    else:
        image = im

    pixels=image.load()
    pixel_converter = partial(rgb_to_xls_hex, image_mode=image.mode)

    # Get the final image size
    size_x, size_y = image.size

    out_wb = Workbook()
    out = out_wb.active

    for y in range(size_y):
        for x in range(size_x):
            cell = out.cell(y+1, x+1)
            rgb = pixels[x, y]
            cell.fill = PatternFill("solid", fgColor=pixel_converter(rgb))

    for col in range(1, size_x+1):
        out.column_dimensions[get_column_letter(col)].width = 3

    out_wb.save(args.xlsx)


if __name__ == "__main__":
    args = handle_arguments()
    convert(args)
